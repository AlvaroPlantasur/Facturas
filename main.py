import os
import psycopg2
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import copy

def main():
    # 1. Obtener credenciales y ruta del archivo
    db_name = os.environ.get('DB_NAME')
    db_user = os.environ.get('DB_USER')
    db_password = os.environ.get('DB_PASSWORD')
    db_host = os.environ.get('DB_HOST')
    db_port = os.environ.get('DB_PORT')
    file_path = os.environ.get('EXCEL_FILE_PATH')

    db_params = {
        'dbname': db_name,
        'user': db_user,
        'password': db_password,
        'host': db_host,
        'port': db_port
    }

    # 2. Definir fechas
    fecha_inicio_str = '2025-01-01'
    fecha_fin = datetime.now().date()
    fecha_fin_str = fecha_fin.strftime('%Y-%m-%d')

    # 3. Consulta SQL
    query = f"""SELECT DISTINCT ON (sm.id)
    sm.invoice_id AS "ID FACTURA",
    sp.name AS "DESCRIPCIÓN",
    sp.internal_number AS "CÓDIGO FACTURA",
    sp.number AS "NÚMERO DEL ASIENTO",
    to_char(sp.date_invoice, 'DD/MM/YYYY') AS "FECHA FACTURA",
    sp.origin AS "DOCUMENTO ORIGEN",
    pp.default_code AS "REFERENCIA PRODUCTO", 
    pp.name AS "NOMBRE", 
    COALESCE(pm.name, '') AS "MARCA",
    s.name AS "SECCION", 
    f.name AS "FAMILIA", 
    sf.name AS "SUBFAMILIA",
    rc.name AS "COMPAÑÍA",
    ssp.name AS "SEDE",
    stp.date_preparado_app AS "FECHA PREPARADO APP",
    (CASE WHEN stp.directo_cliente THEN 'Sí' ELSE 'No' END) AS "CAMIÓN DIRECTO",
    stp.number_of_packages AS "NUMERO DE BULTOS",
    stp.num_pales AS "NUMERO DE PALES",
    sp.portes AS "PORTES",
    sp.portes_cubiertos AS "PORTES CUBIERTOS",
    rp.nombre_comercial AS "CLIENTE",
    rp.vat AS "CIF CLIENTE",
    (CASE 
        WHEN rpa.prov_id IS NOT NULL 
        THEN (SELECT name FROM res_country_provincia WHERE id = rpa.prov_id) 
        ELSE rpa.state_id_2 
    END) AS "PROVINCIA",
    rpa.city AS "CIUDAD",
    (CASE 
        WHEN rpa.cautonoma_id IS NOT NULL 
        THEN (SELECT UPPER(name) FROM res_country_ca WHERE id = rpa.cautonoma_id) 
        ELSE '' 
    END) AS "COMUNIDAD",
    c.name AS "PAÍS",
    EXTRACT(MONTH FROM sp.date_invoice) AS "MES",  -- Cambiado a EXTRACT para número
    EXTRACT(DAY FROM sp.date_invoice) AS "DÍA",   -- Cambiado a EXTRACT para número
    spp.name AS "PREPARADOR",
    sm.peso_arancel AS "PESO",
    SUM(
        CASE 
            WHEN sp.type = 'out_invoice' THEN sm.cantidad_pedida
            WHEN sp.type = 'out_refund' THEN -sm.cantidad_pedida
        END
    ) AS "UNIDADES VENTA",
    SUM(
        CASE 
            WHEN sp.type = 'out_invoice' THEN sm.price_subtotal
            WHEN sp.type = 'out_refund' THEN -sm.price_subtotal
        END
    ) AS "BASE VENTA TOTAL",
    SUM(
        CASE 
            WHEN sp.type = 'out_invoice' THEN sm.margen
            WHEN sp.type = 'out_refund' THEN -sm.margen
        END
    ) AS "MARGEN EUROS",
    SUM(
        CASE 
            WHEN sp.type = 'out_invoice' THEN sm.cantidad_pedida * sm.cost_price_real
            WHEN sp.type = 'out_refund' THEN -sm.cantidad_pedida * sm.cost_price_real
        END
    ) AS "COSTE VENTA TOTAL",
    'S-' || rp.id AS "ID BBSeeds",  -- ID Construdio desde el ID origen
    (CASE 
        WHEN rp.fiscal_position_texto = 'Recargo de Equivalencia' THEN 'Recargo de Equivalencia'
        WHEN rp.fiscal_position_texto = 'Régimen Extracomunitario' THEN 'Régimen Extracomunitario'
        WHEN rp.fiscal_position_texto = 'REGIMEN INTRACOMUNITARIO' THEN 'Régimen Intracomunitario'
        WHEN rp.fiscal_position_texto = 'Régimen Intracomunitario' THEN 'Régimen Intracomunitario'
        WHEN rp.fiscal_position_texto = 'REGIMEN NACIONAL' THEN 'Régimen Nacional'
        ELSE rp.fiscal_position_texto
    END) AS "Tipo Regimen",  -- Nueva columna "Coste Calculado" que es base menos margen
    (SUM(
        CASE 
            WHEN sp.type = 'out_invoice' THEN sm.price_subtotal
            WHEN sp.type = 'out_refund' THEN -sm.price_subtotal
        END
    ) - 
    SUM(
        CASE 
            WHEN sp.type = 'out_invoice' THEN sm.margen
            WHEN sp.type = 'out_refund' THEN -sm.margen
        END
    )) AS "Coste Calculado"
 
FROM account_invoice_line sm
INNER JOIN account_invoice sp ON sp.id = sm.invoice_id
INNER JOIN product_product pp ON sm.product_id = pp.id
INNER JOIN res_partner_address rpa ON sp.address_invoice_id = rpa.id
INNER JOIN res_country c ON c.id = rpa.pais_id
LEFT JOIN stock_picking_invoice_rel spir ON spir.invoice_id = sp.id
LEFT JOIN stock_picking stp ON stp.id = spir.pick_id
LEFT JOIN stock_picking_preparador spp ON spp.id = stp.preparador
LEFT JOIN res_company rc ON rc.id = sp.company_id
LEFT JOIN res_partner rp ON rp.id = sp.partner_id
LEFT JOIN stock_sede_ps ssp ON ssp.id = sp.sede_id
LEFT JOIN product_category s ON s.id = pp.seccion
LEFT JOIN product_category f ON f.id = pp.familia
LEFT JOIN product_category sf ON sf.id = pp.subfamilia
LEFT JOIN product_marca pm ON pm.id = pp.marca
 
WHERE sp.type IN ('out_invoice', 'out_refund') 
AND sp.state IN ('open', 'paid') 
AND sp.journal_id != 11 
AND sp.anticipo = FALSE 
AND pp.default_code NOT LIKE 'XXX%' 
AND sp.obsolescencia = FALSE 
AND rp.nombre_comercial NOT LIKE '%PLANTASUR TRADING%' 
AND rp.nombre_comercial NOT LIKE '%PLANTADUCH%' 
AND sp.date_invoice BETWEEN '{fecha_inicio_str}' AND '{fecha_fin_str}'
 
GROUP BY 
    sm.id, rp.id, sm.company_id, sp.sede_id, sp.date_invoice, 
    pp.seccion, pp.familia, pp.subfamilia, pp.default_code, pp.id, 
    sm.cantidad_pedida, sp.partner_id, rpa.prov_id, rpa.state_id_2, 
    rpa.cautonoma_id, c.name, sp.address_invoice_id, pp.proveedor_id1, 
    sm.price_subtotal, sm.margen, pp.tarifa5, sp.directo_cliente, 
    sp.obsolescencia, sp.anticipo, sp.name, s.name, f.name, sf.name, 
    rc.name, ssp.name, stp.date_preparado_app, stp.directo_cliente, 
    stp.number_of_packages, stp.num_pales, sp.portes, sp.portes_cubiertos, 
    rp.nombre_comercial, spp.name, sp.internal_number, sp.origin, 
    sp.number, rp.vat, rp.fiscal_position_texto, pm.name, rpa.city
 
ORDER BY sm.id, stp.number_of_packages DESC;
    """

    # 4. Ejecutar la consulta
    try:
        with psycopg2.connect(**db_params) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                resultados = cur.fetchall()
                headers = [desc[0] for desc in cur.description]
    except Exception as e:
        print(f"Error al conectar o ejecutar la consulta: {e}")
        sys.exit(1)

    if not resultados:
        print("No se obtuvieron resultados de la consulta.")
        return
    else:
        print(f"Se obtuvieron {len(resultados)} filas de la consulta.")

    # 5. Cargar el archivo Excel
    try:
        book = load_workbook(file_path)
        sheet = book.active
    except FileNotFoundError:
        print(f"No se encontró el archivo '{file_path}'.")
        return

# 6. Borrar todas las filas de datos (manteniendo cabecera)
    sheet.delete_rows(2, sheet.max_row - 1)

# 7. Insertar todos los nuevos registros
    for row in resultados:
        sheet.append(row)

    # 8. Copiar estilo desde fila 2 a nuevas filas
    if sheet.max_row > 2:
        for col in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=2, column=col)
            for row in range(3, sheet.max_row + 1):
                target_cell = sheet.cell(row=row, column=col)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.alignment = copy.copy(source_cell.alignment)


    # 9. Actualizar tabla
    if "Lineas2025" in sheet.tables:
        tabla = sheet.tables["Lineas2025"]
        max_row = sheet.max_row
        max_col = sheet.max_column
        last_col_letter = get_column_letter(max_col)
        new_ref = f"A1:{last_col_letter}{max_row}"
        tabla.ref = new_ref
        print(f"Tabla 'Lineas2025' actualizada a rango: {new_ref}")
    else:
        print("No se encontró la tabla 'Lineas2025'. No se actualizará la referencia.")

    # 10. Guardar archivo
    book.save(file_path)
    print(f"Archivo actualizado correctamente: '{file_path}'.")

if __name__ == '__main__':
    main()
